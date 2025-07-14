from flask import Blueprint, request, jsonify, send_file
from src.models.recipe import db, ShoppingList, ShoppingListItem, MealPlan, Recipe, Ingredient
from datetime import datetime
from collections import defaultdict
import io
import os
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
import openpyxl
from openpyxl.styles import Font, Alignment

shopping_lists_bp = Blueprint('shopping_lists', __name__)

@shopping_lists_bp.route('/shopping-lists', methods=['GET'])
def get_shopping_lists():
    """Get all shopping lists"""
    shopping_lists = ShoppingList.query.order_by(ShoppingList.created_at.desc()).all()
    return jsonify([shopping_list.to_dict() for shopping_list in shopping_lists])

@shopping_lists_bp.route('/shopping-lists/<int:list_id>', methods=['GET'])
def get_shopping_list(list_id):
    """Get a specific shopping list by ID"""
    shopping_list = ShoppingList.query.get_or_404(list_id)
    return jsonify(shopping_list.to_dict())

@shopping_lists_bp.route('/shopping-lists', methods=['POST'])
def create_shopping_list():
    """Create a new shopping list"""
    data = request.get_json()
    
    if not data or 'name' not in data:
        return jsonify({'error': 'Shopping list name is required'}), 400
    
    shopping_list = ShoppingList(name=data['name'])
    
    try:
        db.session.add(shopping_list)
        db.session.flush()  # Get the shopping list ID
        
        # Add items if provided
        if 'items' in data and data['items']:
            for item_data in data['items']:
                if 'ingredient_name' in item_data and 'quantity' in item_data and 'unit' in item_data:
                    item = ShoppingListItem(
                        shopping_list_id=shopping_list.id,
                        ingredient_name=item_data['ingredient_name'],
                        quantity=float(item_data['quantity']),
                        unit=item_data['unit'],
                        checked=item_data.get('checked', False)
                    )
                    db.session.add(item)
        
        db.session.commit()
        return jsonify(shopping_list.to_dict()), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': 'Failed to create shopping list'}), 500

@shopping_lists_bp.route('/shopping-lists/<int:list_id>', methods=['PUT'])
def update_shopping_list(list_id):
    """Update an existing shopping list"""
    shopping_list = ShoppingList.query.get_or_404(list_id)
    data = request.get_json()
    
    if not data:
        return jsonify({'error': 'No data provided'}), 400
    
    if 'name' in data:
        shopping_list.name = data['name']
    
    # Update items if provided
    if 'items' in data:
        # Remove existing items
        ShoppingListItem.query.filter_by(shopping_list_id=list_id).delete()
        
        # Add new items
        for item_data in data['items']:
            if 'ingredient_name' in item_data and 'quantity' in item_data and 'unit' in item_data:
                item = ShoppingListItem(
                    shopping_list_id=list_id,
                    ingredient_name=item_data['ingredient_name'],
                    quantity=float(item_data['quantity']),
                    unit=item_data['unit'],
                    checked=item_data.get('checked', False)
                )
                db.session.add(item)
    
    try:
        db.session.commit()
        return jsonify(shopping_list.to_dict())
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': 'Failed to update shopping list'}), 500

@shopping_lists_bp.route('/shopping-lists/<int:list_id>', methods=['DELETE'])
def delete_shopping_list(list_id):
    """Delete a shopping list"""
    shopping_list = ShoppingList.query.get_or_404(list_id)
    
    try:
        db.session.delete(shopping_list)
        db.session.commit()
        return jsonify({'message': 'Shopping list deleted successfully'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': 'Failed to delete shopping list'}), 500

@shopping_lists_bp.route('/shopping-lists/generate', methods=['POST'])
def generate_shopping_list():
    """Generate a shopping list from meal plans"""
    data = request.get_json()
    
    if not data or 'meal_plan_ids' not in data or 'name' not in data:
        return jsonify({'error': 'Meal plan IDs and shopping list name are required'}), 400
    
    meal_plan_ids = data['meal_plan_ids']
    if not meal_plan_ids:
        return jsonify({'error': 'At least one meal plan ID is required'}), 400
    
    # Get meal plans
    meal_plans = MealPlan.query.filter(MealPlan.id.in_(meal_plan_ids)).all()
    if len(meal_plans) != len(meal_plan_ids):
        return jsonify({'error': 'One or more meal plan IDs are invalid'}), 400
    
    # Aggregate ingredients
    ingredient_totals = defaultdict(lambda: {'quantity': 0, 'unit': '', 'recipes': []})
    
    for meal_plan in meal_plans:
        recipe = meal_plan.recipe
        scaling_factor = meal_plan.people_count / recipe.servings
        
        for ingredient in recipe.ingredients:
            key = ingredient.name.lower().strip()
            adjusted_quantity = ingredient.quantity * scaling_factor
            
            # If this is the first time we see this ingredient, set the unit
            if ingredient_totals[key]['unit'] == '':
                ingredient_totals[key]['unit'] = ingredient.unit
            
            # Only add quantities if units match (simple approach)
            if ingredient_totals[key]['unit'] == ingredient.unit:
                ingredient_totals[key]['quantity'] += adjusted_quantity
            else:
                # Create separate entry for different units
                key_with_unit = f"{key}_{ingredient.unit}"
                ingredient_totals[key_with_unit]['quantity'] += adjusted_quantity
                ingredient_totals[key_with_unit]['unit'] = ingredient.unit
                ingredient_totals[key_with_unit]['recipes'].append(recipe.name)
                continue
            
            ingredient_totals[key]['recipes'].append(recipe.name)
    
    # Create shopping list
    shopping_list = ShoppingList(name=data['name'])
    
    try:
        db.session.add(shopping_list)
        db.session.flush()
        
        # Add aggregated ingredients as shopping list items
        for ingredient_name, details in ingredient_totals.items():
            # Clean up ingredient name (remove unit suffix if added)
            clean_name = ingredient_name.split('_')[0] if '_' in ingredient_name else ingredient_name
            clean_name = clean_name.title()  # Capitalize properly
            
            item = ShoppingListItem(
                shopping_list_id=shopping_list.id,
                ingredient_name=clean_name,
                quantity=round(details['quantity'], 2),
                unit=details['unit'],
                checked=False
            )
            db.session.add(item)
        
        db.session.commit()
        return jsonify(shopping_list.to_dict()), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': 'Failed to generate shopping list'}), 500

@shopping_lists_bp.route('/shopping-lists/<int:list_id>/items/<int:item_id>/toggle', methods=['PUT'])
def toggle_shopping_item(list_id, item_id):
    """Toggle the checked status of a shopping list item"""
    item = ShoppingListItem.query.filter_by(
        id=item_id, 
        shopping_list_id=list_id
    ).first_or_404()
    
    item.checked = not item.checked
    
    try:
        db.session.commit()
        return jsonify(item.to_dict())
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': 'Failed to update item'}), 500

@shopping_lists_bp.route('/shopping-lists/<int:list_id>/export/pdf', methods=['GET'])
def export_shopping_list_pdf(list_id):
    """Export shopping list as PDF"""
    shopping_list = ShoppingList.query.get_or_404(list_id)
    
    # Create PDF in memory
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    story = []
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1  # Center alignment
    )
    story.append(Paragraph(f"Shopping List: {shopping_list.name}", title_style))
    story.append(Spacer(1, 12))
    
    # Date
    date_style = ParagraphStyle(
        'DateStyle',
        parent=styles['Normal'],
        fontSize=10,
        alignment=1
    )
    story.append(Paragraph(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M')}", date_style))
    story.append(Spacer(1, 20))
    
    # Items table
    if shopping_list.items:
        table_data = [['☐', 'Item', 'Quantity', 'Unit']]
        
        for item in shopping_list.items:
            checkbox = '☑' if item.checked else '☐'
            table_data.append([
                checkbox,
                item.ingredient_name,
                str(item.quantity),
                item.unit
            ])
        
        table = Table(table_data, colWidths=[0.5*inch, 3*inch, 1*inch, 1*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(table)
    else:
        story.append(Paragraph("No items in this shopping list.", styles['Normal']))
    
    doc.build(story)
    buffer.seek(0)
    
    # Update exported timestamp
    shopping_list.exported_at = datetime.utcnow()
    db.session.commit()
    
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"shopping_list_{shopping_list.name.replace(' ', '_')}.pdf",
        mimetype='application/pdf'
    )

@shopping_lists_bp.route('/shopping-lists/<int:list_id>/export/excel', methods=['GET'])
def export_shopping_list_excel(list_id):
    """Export shopping list as Excel file"""
    shopping_list = ShoppingList.query.get_or_404(list_id)
    
    # Create Excel workbook in memory
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Shopping List"
    
    # Title
    ws['A1'] = f"Shopping List: {shopping_list.name}"
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:D1')
    
    # Date
    ws['A2'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws['A2'].font = Font(size=10)
    ws.merge_cells('A2:D2')
    
    # Headers
    headers = ['Checked', 'Item', 'Quantity', 'Unit']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Items
    for row, item in enumerate(shopping_list.items, 5):
        ws.cell(row=row, column=1, value='✓' if item.checked else '')
        ws.cell(row=row, column=2, value=item.ingredient_name)
        ws.cell(row=row, column=3, value=item.quantity)
        ws.cell(row=row, column=4, value=item.unit)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Update exported timestamp
    shopping_list.exported_at = datetime.utcnow()
    db.session.commit()
    
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"shopping_list_{shopping_list.name.replace(' ', '_')}.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

